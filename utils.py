from calendar import monthrange
from datetime import datetime
import os
import openpyxl
import pandas as pd

from functions import get_name_month_now, get_count_day_month, get_date_update_file

month = datetime.now().month
# print(month)
name_month = get_name_month_now(month)


f = f"C:/Users/User1/Desktop/расчетки/{name_month}/Суточный рапорт за {name_month.lower()} c трелевкой.xlsx"
# f = "C:/Users/User1/Desktop/расчетки/Декабрь/Суточный рапорт за ноябрь c трелевкой.xlsx"
# f = "C:/Users/User1/Desktop/расчетки/Логистика.xlsx"
# date_today = datetime.today()
# year_today = datetime.now().year
# month_today = datetime.now().month

# количество дней в месяце
count_days = get_count_day_month(month)
# xl_obj = openpyxl.open(f)


# sh = xl_obj.worksheets[0]
#
# print(sh['B2'].value)

ddd = openpyxl.load_workbook(f, data_only=True)
# print(get_plan(ddd, count_days))
d = ddd['2']

# # кубы по графику
# count_graf = int(d['I6'].value)
#
# # кубы по факту
# count_fact = int(d['K6'].value)
#
# # отклонение с начала месяца
# count_fact = count_graf - count_fact

# data = pd.read_excel(f, sheet_name=['1'])
# print(data['Заготовка'])

# c = sh.cell(row=13, column=1)

# for i in sh.iter_cols(min_col=1, max_col=31, values_only=True, max_row=1):
#     print(datetime.date(i[0]))
#     print(date_go)
    # if i[0] == date_go:
    #     print(i)
# print(c.value)
# print(date_go.date())
# print(date_go)



#дата обновления файла
update_time = os.path.getmtime(f)
mtime_readable = datetime.fromtimestamp(update_time)

number_sheet = mtime_readable.day
a = get_date_update_file(f)
print(a)
